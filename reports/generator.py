"""
Excel Report Generator
Produces a multi-sheet, colour-coded .xlsx file with:
  Sheet 1 – SUMMARY        (all stocks, ranked by score)
  Sheet 2 – STRONG_SETUPS  (score ≥ 6, green verdicts)
  Sheet 3 – WATCHLIST      (score 4-5, blue/amber verdicts)
  Sheet 4 – RISK_TABLE     (detailed risk profiles)
  Sheet 5 – METADATA       (run parameters)
"""

import os
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import REPORTS_DIR, DEFAULT_CAPITAL, RISK_PCT
from scoring.scorecard import ScorecardResult


# ── Colour palette ────────────────────────────────────────────────────────────
COLORS = {
    "GREEN":   "C6EFCE",   # Strong bullish
    "BLUE":    "BDD7EE",   # Watch
    "AMBER":   "FFEB9C",   # Retest
    "RED":     "FFC7CE",   # Downtrend
    "GREY":    "D9D9D9",   # Skip
    "HEADER":  "1F4E79",   # Dark navy header
    "SUBHDR":  "2E75B6",   # Mid-blue sub-header
    "MET":     "C6EFCE",
    "PARTIAL": "FFEB9C",
    "NOT MET": "FFC7CE",
    "WHITE":   "FFFFFF",
    "LIGHT":   "F2F2F2",
}

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=9)
TITLE_FONT   = Font(bold=True, size=11, color="1F4E79")

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_row(ws, row: int, values: list, col_start: int = 1):
    for ci, val in enumerate(values, start=col_start):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill   = _fill(COLORS["HEADER"])
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=8, max_w=35):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(max_len + 2, min_w), max_w)


def _results_to_df(results: list[ScorecardResult]) -> pd.DataFrame:
    rows = []
    for r in results:
        rows.append({
            "Symbol":       r.symbol,
            "CMP (₹)":      r.cmp,
            "Score":        f"{r.score}/8",
            "Score_int":    r.score,
            "Verdict":      r.verdict,
            "Trend":        r.trend_state,
            "Gate Trend":   r.gate_trend,
            "Gate Support": r.gate_support,
            "Gate Candle":  r.gate_candle,
            "Gate RR":      r.gate_rr,
            "Gates Met":    r.gates_met,
            "Candle":       r.candle_pattern,
            "Support ₹":    r.support_zone,
            "Resistance ₹": r.resistance_zone,
            "Stop Loss ₹":  r.stop_loss,
            "Target 1 ₹":   r.target1,
            "RR T1":        r.rr_t1,
            "RR Grade":     r.rr_grade,
            "Beta":         r.beta,
            "Avg Vol 20D":  r.avg_vol_20,
            "ATR 14":       r.atr14,
            "Verdict Color":r.verdict_color,
        })
    df = pd.DataFrame(rows).sort_values("Score_int", ascending=False)
    return df


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame, color_col: str = "Verdict Color"):
    ws = wb.create_sheet(sheet_name)
    display_cols = [c for c in df.columns if c not in ("Score_int", "Verdict Color")]
    _header_row(ws, 1, display_cols)

    verdict_color_idx = list(df.columns).index(color_col) if color_col in df.columns else None

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        row_dict   = dict(zip(df.columns, row_data))
        v_color    = row_dict.get("Verdict Color", "GREY")
        row_fill   = _fill(COLORS.get(v_color, COLORS["GREY"]))
        alt_fill   = _fill(COLORS["LIGHT"]) if ri % 2 == 0 else _fill(COLORS["WHITE"])

        for ci, col in enumerate(display_cols, start=1):
            val  = row_dict.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == "Verdict":
                cell.fill = row_fill
            elif col in ("Gate Trend", "Gate Support", "Gate Candle", "Gate RR"):
                gate_color = {"YES": "C6EFCE", "PARTIAL": "FFEB9C", "NO": "FFC7CE"}.get(str(val), "FFFFFF")
                cell.fill  = _fill(gate_color)
            elif col == "Score":
                score_int = row_dict.get("Score_int", 0)
                cell.fill = _fill(
                    "C6EFCE" if score_int >= 6
                    else "FFEB9C" if score_int >= 4
                    else "FFC7CE"
                )
            else:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    _auto_width(ws)
    return ws


def _write_scorecard_detail(wb, results: list[ScorecardResult]):
    """Sheet with full 8-criterion breakdown for top 50 stocks."""
    ws = wb.create_sheet("SCORECARD_DETAIL")
    headers = ["Symbol", "CMP", "Score", "Criterion", "Status", "Note"]
    _header_row(ws, 1, headers)

    row = 2
    for r in results[:50]:   # top 50 by score
        items = r.get("score_items", [])
        for item in items:
            ws.cell(row, 1, r['symbol']).font  = BODY_FONT
            ws.cell(row, 2, r['ltp']).font     = BODY_FONT
            ws.cell(row, 3, r['score']).font   = BODY_FONT
            ws.cell(row, 4, item.criterion).font = BODY_FONT
            status_cell = ws.cell(row, 5, item.status)
            status_cell.fill = _fill(COLORS.get(item.status, "FFFFFF"))
            status_cell.font = BODY_FONT
            ws.cell(row, 6, item.note).font = BODY_FONT
            for ci in range(1, 7):
                ws.cell(row, ci).border    = THIN_BORDER
                ws.cell(row, ci).alignment = Alignment(vertical="center", wrap_text=True)
            row += 1
        row += 1   # blank separator between stocks

    ws.freeze_panes = "A2"
    ws.column_dimensions["F"].width = 60
    _auto_width(ws)


def _write_metadata(wb, total_scanned: int, elapsed: float):
    ws = wb.create_sheet("METADATA")
    info = [
        ("Report Generated",    datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")),
        ("Analysis Framework",  "Institutional Swing Analyzer v2.0"),
        ("Universe",            "NSE Equities (India)"),
        ("Timeframe",           "Daily (1D)"),
        ("Lookback Period",     "6 Months"),
        ("Stocks Scanned",      total_scanned),
        ("Processing Time (s)", round(elapsed, 1)),
        ("Capital Assumed",     f"₹{DEFAULT_CAPITAL:,.0f}"),
        ("Risk Per Trade",      f"{RISK_PCT*100}%"),
        ("SEBI Disclaimer",     "This report is for educational purposes only. "
                                 "Data may have a 30-day lag for unregistered entities. "
                                 "Not investment advice. Consult a SEBI-registered advisor."),
    ]
    ws.cell(1, 1, "Institutional Swing Analyzer – Run Metadata").font = TITLE_FONT
    for ri, (k, v) in enumerate(info, start=3):
        ws.cell(ri, 1, k).font  = Font(bold=True, size=9)
        ws.cell(ri, 2, str(v)).font = BODY_FONT
    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────
def generate_report(
    results: list[dict],
    elapsed: float = 0.0,
    output_filename: str = "institutional_swing_analysis.xlsx",
) -> str:
    """
    Writes the full Excel report and returns the file path.
    Accepts list of dictionaries from scan_engine.py.
    """
    output_path = os.path.join(REPORTS_DIR, output_filename)
    os.makedirs(REPORTS_DIR, exist_ok=True)

    # Convert dictionaries back to ScorecardResult objects for processing
    # (or just adapt _results_to_df to handle dicts)
    
    rows = []
    for r in results:
        rows.append({
            "Symbol":       r["symbol"],
            "CMP (₹)":      r["ltp"],
            "Score":        f"{r['score']}/8",
            "Score_int":    r["score"],
            "Verdict":      r["verdict"],
            "Trend":        r["trend_state"],
            "Gate Trend":   r.get("gate_trend", "NO"),
            "Gate Support": r.get("gate_support", "NO"),
            "Gate Candle":  r.get("gate_candle", "NO"),
            "Gate RR":      r.get("gate_rr", "NO"),
            "Gates Met":    r.get("gates_met", 0),
            "Candle":       r["pattern"],
            "Support ₹":    r.get("support_zone"),
            "Resistance ₹": r.get("target"),
            "Stop Loss ₹":  r["sl"],
            "Target 1 ₹":   r["target"],
            "RR T1":        r["rr"],
            "RR Grade":     r.get("rr_grade", ""),
            "Beta":         r["beta"],
            "Avg Vol 20D":  r.get("avg_vol_20", 0),
            "ATR 14":       r.get("atr14", 0),
            "Verdict Color":r["verdict_color"],
        })
    df = pd.DataFrame(rows).sort_values("Score_int", ascending=False)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Sheet 1: Summary (all) ────────────────────────────────────────────────
    _write_sheet(wb, "SUMMARY", df)

    # ── Sheet 2: Strong setups ────────────────────────────────────────────────
    strong = df[df["Score_int"] >= 6].copy()
    _write_sheet(wb, "STRONG_SETUPS", strong)

    # ── Sheet 3: Watchlist ────────────────────────────────────────────────────
    watch = df[(df["Score_int"] >= 4) & (df["Score_int"] < 6)].copy()
    _write_sheet(wb, "WATCHLIST", watch)

    # ── Sheet 4: Downtrends (avoid) ───────────────────────────────────────────
    down = df[df["Trend"] == "DOWNTREND"].copy()
    _write_sheet(wb, "DOWNTRENDS_AVOID", down)

    # ── Sheet 5: Full scorecard detail ───────────────────────────────────────
    sorted_results = sorted(results, key=lambda r: r['score'], reverse=True)
    _write_scorecard_detail(wb, sorted_results)

    # ── Sheet 6: Metadata ─────────────────────────────────────────────────────
    _write_metadata(wb, len(results), elapsed)

    wb.save(output_path)
    return output_path
